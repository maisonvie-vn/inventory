import sys
import json
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

print("=== SEARCHING IN DB.JSON ===")
with open(r'd:\Invenroty\maison-vie-crm\src\data\db.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

for category in ['sales', 'recipes']:
    items = data.get(category, [])
    if isinstance(items, dict):
        items = list(items.values())
    for item in items:
        name = item.get('name', '')
        code = item.get('code', '') or item.get('id', '')
        if any(x in str(name).lower() or x in str(code).lower() for x in ['set', '350', '370', '470']):
            print(f"[{category}] Code: {code} | Name: {name} | Price: {item.get('price', 0)}")

print("\n=== SEARCHING IN MAISON_VIE_v6_0_PRO.xlsx ===")
try:
    wb = openpyxl.load_workbook(r'D:\Invenroty\MAISON_VIE_v6_0_PRO.xlsx', read_only=True)
    for sname in wb.sheetnames:
        ws = wb[sname]
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            row_str = ' '.join([str(x) for x in row if x is not None]).lower()
            if any(x in row_str for x in ['set', '350', '370', '470']):
                print(f"[{sname}] Row {idx+1}: {row[:6]}")
except Exception as e:
    print("Error reading Excel:", e)
