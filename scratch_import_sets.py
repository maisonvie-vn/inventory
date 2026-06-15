import sys
import json
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r'D:\Invenroty\MAISON_VIE_v6_0_PRO.xlsx'
db_json_path = r'd:\Invenroty\maison-vie-crm\src\data\db.json'
mock_data_path = r'd:\Invenroty\maison-vie-crm\src\data\mockData.ts'

wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['RECIPE_BEP']

# Read all ingredients from db.json to get their correct prices/details
with open(db_json_path, 'r', encoding='utf-8') as f:
    db_data = json.load(f)

ingredients_list = db_data.get('ingredients', [])
ing_prices = {ing['id']: ing['price'] for ing in ingredients_list}
ing_units = {ing['id']: ing['unit'] for ing in ingredients_list}

# We will parse RECIPE_BEP
# Format of columns: 0: Mã Món, 1: Tên, 2: Giá, 3: Mã NVL, 4: Tên NVL, 5: DL/p, 6: Đơn giá, 7: Cost, 8: FC%
recipes_from_excel = {}
current_recipe = None

for row_idx in range(2, ws.max_row + 1):
    vals = [ws.cell(row_idx, col_idx).value for col_idx in range(1, 10)]
    menu_code = vals[0]
    menu_name = vals[1]
    menu_price = vals[2]
    ing_code = vals[3]
    ing_name = vals[4]
    qty_net = vals[5]
    
    # Check if this is a new recipe start
    if menu_code is not None and str(menu_code).strip() != '':
        code = str(menu_code).strip()
        name = str(menu_name).strip() if menu_name else f"Món {code}"
        price = float(menu_price) if menu_price is not None else 0.0
        
        current_recipe = {
            "code": code,
            "name": name,
            "course": "Bếp",
            "category": "Set Menu" if "set" in code.lower() or code.startswith('T') else "Món chính",
            "menu_role": "Set Menu" if "set" in code.lower() or code.startswith('T') else "À la carte",
            "price": price,
            "ingredients": [],
            "method": [],
            "notes": ""
        }
        recipes_from_excel[code] = current_recipe

    if current_recipe and ing_code and str(ing_code).strip() != '—' and str(ing_code).strip() != '':
        i_code = str(ing_code).strip()
        i_qty_net = float(qty_net) if qty_net is not None else 0.0
        
        # Look up price and unit from our master list
        unit_price = ing_prices.get(i_code, float(vals[6]) if vals[6] is not None else 0.0)
        unit = ing_units.get(i_code, 'kg')
        yield_pct = 0.85 # default yield
        
        # try to get yield rate from ingredient if exists
        ing_obj = next((x for x in ingredients_list if x['id'] == i_code), None)
        if ing_obj:
            yield_pct = ing_obj.get('yield_rate', 0.85)
            
        qty_eff = i_qty_net / yield_pct if yield_pct > 0 else i_qty_net
        line_cost = qty_eff * unit_price
        
        current_recipe["ingredients"].append({
            "ing_id": i_code,
            "qty_net": i_qty_net,
            "unit": unit,
            "yield_pct": yield_pct,
            "qty_eff": qty_eff,
            "unit_price": unit_price,
            "line_cost": line_cost
        })

print(f"Parsed {len(recipes_from_excel)} recipes from Excel.")

# Select only the recipes starting with SET or T (for Tour sets)
set_recipes = {code: r for code, r in recipes_from_excel.items() if code.startswith('SET') or (code.startswith('T') and any(char.isdigit() for char in code))}
print(f"Found {len(set_recipes)} Set Menu recipes in Excel: {list(set_recipes.keys())}")

# Update db.json recipes
existing_recipes = db_data.get('recipes', {})
if isinstance(existing_recipes, list):
    # convert to dict if it was a list
    existing_recipes = {r['code']: r for r in existing_recipes}

# Add or overwrite set recipes in db.json
for code, r in set_recipes.items():
    existing_recipes[code] = r

db_data['recipes'] = existing_recipes

# Save back to db.json
with open(db_json_path, 'w', encoding='utf-8') as f:
    json.dump(db_data, f, ensure_ascii=False, indent=2)
print("Updated db.json with Set Menu recipes!")

# Now update POS_MAPPING and SET_MENU_DEFINITIONS in mockData.ts to include these Set menus
# We will check if there are sales records in Excel for these Set menus
sales_ws = wb['BEP_T06'] # Let's check June sales
excel_sales_codes = set()
for r in range(6, sales_ws.max_row+1):
    c_code = sales_ws.cell(r, 1).value
    if c_code:
        excel_sales_codes.add(str(c_code).strip())

print("Sales codes in BEP_T06:", sorted(list(excel_sales_codes)))

# We need to write update to mockData.ts
# Let's inspect mockData.ts structure first to see how we can insert these new POS mappings
